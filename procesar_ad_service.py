"""
ElectroAlem - Microservicio de procesamiento de archivos AD
Extrae métricas estandarizadas del XLSX para el flujo n8n.

Deploy: gunicorn en Railway
Uso: POST /procesar-ad con multipart/form-data (file + filename + historial opcional)
     POST /generar-ad  con multipart/form-data (archivo_sistema + archivo_anterior opcional)
"""

from flask import Flask, request, jsonify, send_file
import pandas as pd
import re
import json
import io
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINT EXISTENTE: /procesar-ad
# ══════════════════════════════════════════════════════════════════════════════

def extraer_fecha_desde_nombre(filename: str) -> str:
    """
    Extrae fecha del nombre del archivo.
    Soporta: AD_DD_MM_YYYY.xlsx y AD DD_MM_YYYY.xlsx (con espacio)
    """
    normalizado = filename.replace(' ', '_')
    match = re.search(r'AD_(\d{2})_(\d{2})_(\d{4})', normalizado)
    if match:
        d, m, y = match.groups()
        return f"{y}-{m}-{d}"
    return datetime.now().strftime("%Y-%m-%d")


def procesar_ad(file_bytes: bytes, filename: str) -> dict:
    fecha_ad = extraer_fecha_desde_nombre(filename)

    # ── Tabla Dinam: deuda por cliente y tramos ──────────────────────────────
    tdf = pd.read_excel(BytesIO(file_bytes), sheet_name='Tabla Dinam', header=None)
    tdf.columns = tdf.iloc[3]
    tdf = tdf.iloc[4:].reset_index(drop=True)

    for col in ['>61', '46-60', '31-45', '16-30', '1-15', 'Total general']:
        tdf[col] = pd.to_numeric(tdf[col], errors='coerce').fillna(0)

    tdf['Código'] = tdf['Código'].astype(str)

    # Excluir fila de totales
    tdf = tdf[~tdf['Código'].str.contains('Total', na=False)]

    morosos = tdf[tdf['Total general'] > 0]
    total_cuentas = len(tdf[tdf['Total general'] != 0])
    cuentas_morosas = morosos['Código'].nunique()

    mora_total = morosos['Total general'].sum()
    mora_mayor_61 = morosos['>61'].sum()
    mora_46_60 = morosos['46-60'].sum()
    mora_31_45 = morosos['31-45'].sum()
    mora_16_30 = morosos['16-30'].sum()
    mora_1_15 = morosos['1-15'].sum()

    # Mora real = 15 a 61 días
    mora_real_15_61 = mora_16_30 + mora_31_45 + mora_46_60

    # Ticket promedio
    ticket_promedio = mora_total / cuentas_morosas if cuentas_morosas > 0 else 0

    # Top 5 deudores
    top5 = morosos.nlargest(5, 'Total general')[['Código', 'Razón Social', 'Total general']]
    top5_list = top5.to_dict(orient='records')

    # ── Trabajando: cobranza semanal y evolución ──────────────────────────────
    wdf = pd.read_excel(BytesIO(file_bytes), sheet_name='Trabajando', header=None)

    try:
        cob_prev = float(wdf.iloc[7, 16]) if pd.notna(wdf.iloc[7, 16]) else 0
        cob_curr = float(wdf.iloc[7, 17]) if pd.notna(wdf.iloc[7, 17]) else 0
        evo_cobranza = float(wdf.iloc[7, 18]) if pd.notna(wdf.iloc[7, 18]) else 0
    except (IndexError, ValueError, TypeError):
        cob_prev, cob_curr, evo_cobranza = 0, 0, 0

    # Performance = cobranza / mora real (%)
    performance_pct = (cob_curr / mora_real_15_61 * 100) if mora_real_15_61 > 0 else 0

    # ── Cantidades por tramo ──────────────────────────────────────────────────
    cant_mayor_61 = int((morosos['>61'] > 0).sum())
    cant_46_60 = int((morosos['46-60'] > 0).sum())
    cant_31_45 = int((morosos['31-45'] > 0).sum())
    cant_16_30 = int((morosos['16-30'] > 0).sum())
    cant_1_15 = int((morosos['1-15'] > 0).sum())

    return {
        "fecha_ad": fecha_ad,
        "filename": filename,
        "total_cuentas": total_cuentas,
        "cuentas_morosas": cuentas_morosas,
        "mora_total": round(mora_total, 2),
        "mora_real_15_61": round(mora_real_15_61, 2),
        "mora_mayor_61": round(mora_mayor_61, 2),
        "mora_46_60": round(mora_46_60, 2),
        "mora_31_45": round(mora_31_45, 2),
        "mora_16_30": round(mora_16_30, 2),
        "mora_1_15": round(mora_1_15, 2),
        "cobranza_semanal": round(cob_curr, 2),
        "cobranza_semana_anterior": round(cob_prev, 2),
        "performance_cobranza_pct": round(performance_pct, 2),
        "ticket_promedio": round(ticket_promedio, 2),
        "evolucion_cobranza": round(evo_cobranza, 6),
        "cant_mayor_61": cant_mayor_61,
        "cant_46_60": cant_46_60,
        "cant_31_45": cant_31_45,
        "cant_16_30": cant_16_30,
        "cant_1_15": cant_1_15,
        "top5_deudores_json": json.dumps(top5_list, ensure_ascii=False),
        "codigos_morosos": list(morosos['Código'].unique()),
    }


def calcular_reincidentes(codigos_actuales: list, historial_rows: list) -> dict:
    """
    Calcula reincidentes y clientes que pagaron comparando con el historial.
    - Reincidentes: morosos en el AD actual que TAMBIÉN estaban en TODAS las semanas anteriores
    - Pagaron: estaban en la semana más antigua del historial y YA NO están en el actual
    """
    if not historial_rows:
        return {"reincidentes_count": 0, "pagaron_count": 0}

    set_actual = set(str(c) for c in codigos_actuales)

    # Reconstruir sets históricos desde el campo codigos_morosos del sheet
    sets_historicos = []
    for row in historial_rows:
        codigos_str = row.get('codigos_morosos', '[]')
        try:
            parsed = json.loads(codigos_str) if isinstance(codigos_str, str) else codigos_str
            if isinstance(parsed, list) and len(parsed) > 0:
                sets_historicos.append(set(str(c) for c in parsed))
        except Exception:
            pass

    if not sets_historicos:
        return {"reincidentes_count": 0, "pagaron_count": 0}

    # Reincidentes: presentes en el actual Y en TODAS las semanas históricas
    reincidentes = set_actual.copy()
    for s in sets_historicos:
        reincidentes &= s

    # Pagaron: estaban en la semana más antigua y no están en el actual
    set_mas_antiguo = sets_historicos[0]
    pagaron = set_mas_antiguo - set_actual

    return {
        "reincidentes_count": len(reincidentes),
        "pagaron_count": len(pagaron),
    }


@app.route('/procesar-ad', methods=['POST'])
def endpoint_procesar():
    if 'file' not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400

    f = request.files['file']
    filename = request.form.get('filename', f.filename)
    file_bytes = f.read()

    try:
        resultado = procesar_ad(file_bytes, filename)

        # Calcular reincidentes con historial si viene en el request
        historial_json = request.form.get('historial', '[]')
        try:
            historial = json.loads(historial_json)
        except Exception:
            historial = []

        reincidentes = calcular_reincidentes(resultado['codigos_morosos'], historial)
        resultado.update(reincidentes)

        return jsonify(resultado)
    except Exception as e:
        return jsonify({"error": str(e), "filename": filename}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ENDPOINT NUEVO: /generar-ad
# Recibe el archivo crudo del sistema y genera el archivo completo de gestoras
# ══════════════════════════════════════════════════════════════════════════════

# ── Helpers ──────────────────────────────────────────────────────────────────

def _tramo(dias):
    """Clasifica días vencidos en tramo de mora."""
    if pd.isna(dias):
        return None
    d = int(dias)
    if d > 61:   return ">61"
    if d > 45:   return "46-60"
    if d > 30:   return "31-45"
    if d > 15:   return "16-30"
    if d > 0:    return "1-15"
    return None


def _header_style(ws, row, cols, bg="1F4E79", fg="FFFFFF"):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=fg, size=10)
    aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill, c.font, c.alignment = fill, font, aln


def _auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


def _write_df(ws, df, start_row=1):
    for ci, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=ci, value=col_name)
    for ri, row in enumerate(df.itertuples(index=False), start_row + 1):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci)
            if isinstance(val, float) and pd.isna(val):
                cell.value = None
            elif hasattr(val, 'item'):
                cell.value = val.item()
            else:
                cell.value = val


def _semana_range(fecha_ref, offset_weeks=0):
    lunes  = fecha_ref - timedelta(days=fecha_ref.weekday()) + timedelta(weeks=offset_weeks)
    domingo = lunes + timedelta(days=6)
    return lunes, domingo


def _semana_labels(fecha_ref):
    ini_ant, fin_ant = _semana_range(fecha_ref, -1)
    ini_act, fin_act = _semana_range(fecha_ref,  0)
    fmt = "%d/%m"
    return {
        "anterior": f"Cobranzas {ini_ant.strftime(fmt)} al  {fin_ant.strftime(fmt)}",
        "actual":   f"Cobranzas {ini_act.strftime(fmt)} al  {fin_act.strftime(fmt)}",
    }


# ── Transformación principal ──────────────────────────────────────────────────

def _transformar(sistema_bytes, anterior_bytes=None, fecha_ref=None):
    if fecha_ref is None:
        # La fecha de referencia es el lunes de la semana siguiente al archivo.
        # Esto replica la lógica manual: los días se calculan desde el próximo lunes.
        hoy = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        dias_hasta_lunes = (7 - hoy.weekday()) % 7
        if dias_hasta_lunes == 0:
            dias_hasta_lunes = 7
        fecha_ref = hoy + timedelta(days=dias_hasta_lunes)

    # 1. Leer archivo del sistema
    df = pd.read_excel(io.BytesIO(sistema_bytes), sheet_name=0, header=1)
    df.columns = df.columns.str.strip()
    df = df.dropna(subset=["Código"]).reset_index(drop=True)
    df["Código"] = df["Código"].astype(int)

    # 2. Leer tablas de referencia del archivo anterior (gestoras semana previa)
    df_listado   = pd.DataFrame()
    df_portales  = pd.DataFrame()
    df_vend      = pd.DataFrame()
    df_trab_prev = pd.DataFrame()

    if anterior_bytes:
        ref = io.BytesIO(anterior_bytes)
        try:
            df_listado = pd.read_excel(ref, sheet_name="Listado", header=1)
            df_listado.columns = df_listado.columns.str.strip()
            df_listado["Cód."] = pd.to_numeric(df_listado["Cód."], errors="coerce")
            df_listado = df_listado.dropna(subset=["Cód."]).copy()
            df_listado["Cód."] = df_listado["Cód."].astype(int)
        except Exception:
            pass

        ref.seek(0)
        try:
            df_portales = pd.read_excel(ref, sheet_name="Portales", header=1)
            df_portales.columns = df_portales.columns.str.strip()
            df_portales["CODIGO"] = pd.to_numeric(df_portales["CODIGO"], errors="coerce").astype("Int64")
        except Exception:
            pass

        ref.seek(0)
        try:
            df_vend = pd.read_excel(ref, sheet_name="Vendedores (3)", header=1)
            df_vend.columns = df_vend.columns.str.strip()
            df_vend["N°"] = pd.to_numeric(df_vend["N°"], errors="coerce").astype("Int64")
        except Exception:
            pass

        ref.seek(0)
        try:
            df_trab_prev = pd.read_excel(ref, sheet_name="Trabajando", header=9)
            df_trab_prev.columns = df_trab_prev.columns.str.strip()
            df_trab_prev = df_trab_prev.rename(columns={df_trab_prev.columns[0]: "Cod"})
            df_trab_prev["Cod"] = pd.to_numeric(df_trab_prev["Cod"], errors="coerce").astype("Int64")
            df_trab_prev = df_trab_prev.dropna(subset=["Cod"])
        except Exception:
            pass

    # 3. Calcular días vencidos y tramo
    df["Días"]  = (fecha_ref - pd.to_datetime(df["Fecha Vto"])).dt.days
    df["Tramo"] = df["Días"].apply(_tramo)

    # 4. Enriquecer con Listado
    if not df_listado.empty:
        lkp = df_listado.set_index("Cód.")[["Lím,Crédito", "Clasificación", "Categ."]].rename(columns={"Lím,Crédito": "Límite"})
        df  = df.join(lkp, on="Código", how="left")
    else:
        df["Límite"] = df["Clasificación"] = df["Categ."] = None

    # 5. Enriquecer con Vendedores
    if not df_vend.empty:
        lkp = df_vend.set_index("N°")[["Nombre", "C.Costos"]].rename(columns={"Nombre": "Vendedores", "C.Costos": "C.Co."})
        df  = df.join(lkp, on="Vend", how="left")
    else:
        df["Vendedores"] = df["C.Co."] = None

    # 6. Enriquecer con Portales
    if not df_portales.empty:
        df["P"] = df["Código"].map(df_portales.set_index("CODIGO")["Portal"])
    else:
        df["P"] = None

    # 7. Base Deuda
    cols_bd = ["Razón Social","Código","Vend","Fecha Doc.","Documento","L","Pref","N°",
               "Fecha Vto","Imp.Original","Imp Cancelado","Saldo Documento","Saldo Total",
               "Ej.Cta","Cód Vta","Límite","Clasificación","Vendedores","Días","P","C.Co.","Categ."]
    df_base_deuda = df[[c for c in cols_bd if c in df.columns]].copy()

    # 8. Tabla dinámica por cliente y tramo
    # Incluir todos los documentos con saldo distinto de cero.
    # Los no vencidos (días <= 0) se agrupan en "1-15" para que aparezcan en la tabla.
    tramos = [">61", "46-60", "31-45", "16-30", "1-15"]
    # Pivot usa Saldo Documento (saldo individual por factura, no acumulativo)
    # Solo filas con saldo distinto de cero y tramo calculado (documentos vencidos)
    df_pivot_src = df[(df["Saldo Documento"] != 0) & df["Tramo"].notna()].copy()
    pivot = df_pivot_src.pivot_table(
        index=["Código","Razón Social","C.Co.","Ej.Cta","P","Límite","Clasificación","Categ."],
        columns="Tramo", values="Saldo Documento", aggfunc="sum"
    ).reset_index()
    for t in tramos:
        if t not in pivot.columns:
            pivot[t] = None
    pivot["Total general"] = pivot[[t for t in tramos if t in pivot.columns]].sum(axis=1, min_count=1)
    df_pivot = pivot[["Código","Razón Social","C.Co.","Ej.Cta","P","Límite","Clasificación","Categ."] + tramos + ["Total general"]]

    # 9. Hoja Trabajando
    labels = _semana_labels(fecha_ref)
    df_trab = df_pivot.copy().rename(columns={"Código": "Cod"})
    for col in ["Fecha1","Acción1","Fecha2","Acción2","Interv Ej cta"]:
        df_trab[col] = None
    obs_nueva    = f"OBS. SEMANA_{fecha_ref.strftime('%d_%m')}"
    obs_anterior = None
    df_trab[obs_nueva] = None

    # Preservar observaciones y acciones de la semana anterior
    if not df_trab_prev.empty:
        prev_obs  = [c for c in df_trab_prev.columns if c.startswith("OBS.")]
        prev_man  = [c for c in ["Fecha1","Acción1","Fecha2","Acción2","Interv Ej cta"] if c in df_trab_prev.columns]
        lkp_prev  = df_trab_prev[["Cod"] + prev_man + prev_obs].drop_duplicates(subset=["Cod"])
        df_trab   = df_trab.merge(lkp_prev, on="Cod", how="left", suffixes=("", "_prev"))
        for col in prev_man:
            if col + "_prev" in df_trab.columns:
                df_trab[col] = df_trab[col].fillna(df_trab[col + "_prev"])
                df_trab.drop(columns=[col + "_prev"], inplace=True)
        if prev_obs:
            obs_anterior = prev_obs[0]
            for c in [c for c in df_trab.columns if c.endswith("_prev")]:
                df_trab.drop(columns=[c], inplace=True)

    # 10. Cobranzas semanales (recibos filtrados por rango de fecha)
    df_recibos = df[df["Documento"].str.contains("RECIBO", na=False, case=False)].copy()

    def _cob(ini, fin):
        mask = (df_recibos["Fecha Doc."] >= ini) & (df_recibos["Fecha Doc."] <= fin)
        return df_recibos[mask][["Código","Razón Social","Fecha Doc.","Documento","Saldo Total"]].rename(
            columns={"Razón Social":"Razon  Social","Fecha Doc.":"Fecha","Saldo Total":"Importe"})

    ini_ant, fin_ant = _semana_range(fecha_ref, -1)
    ini_act, fin_act = _semana_range(fecha_ref,  0)
    df_cob_ant = _cob(ini_ant, fin_ant)
    df_cob_act = _cob(ini_act, fin_act)

    df_trab[labels["anterior"]] = df_trab["Cod"].map(df_cob_ant.groupby("Código")["Importe"].sum())
    df_trab[labels["actual"]]   = df_trab["Cod"].map(df_cob_act.groupby("Código")["Importe"].sum())

    # 11. Escribir XLSX
    wb = Workbook()
    wb.remove(wb.active)

    # Hoja: Tabla (3)
    ws = wb.create_sheet("Tabla (3)")
    tabla3 = pd.DataFrame({
        "Respuestas": ["1- Mail masivo","2- Mail","3- Llamado","4- Whatsapp","5- Visita",
                       "6- En gestión","7- Pago para retirar","8- Canceló","9- Pago parcial",
                       "10-Promesa de pago","11- Otro","12-KOMMO"],
        "Vendedor":   ["ANIBAL MORENO","PIVA HECTOR","OLMOS","CROSA JOSE","CARAFFA",
                       "MALDONADO MARCELO","SANCHEZ OSVALDO","JORGE MAJUL - CTAS ESPECIALES",
                       "JORGE FERRINI","TITO","QUINTEROS PABLO","ARAYA RICARDO"]
    })
    _write_df(ws, tabla3); _header_style(ws, 1, 2); _auto_width(ws)

    # Hoja: Vendedores (3)
    ws = wb.create_sheet("Vendedores (3)")
    if not df_vend.empty:
        _write_df(ws, df_vend); _header_style(ws, 1, len(df_vend.columns)); _auto_width(ws)

    # Hoja: Portales
    ws = wb.create_sheet("Portales")
    if not df_portales.empty:
        _write_df(ws, df_portales); _header_style(ws, 1, len(df_portales.columns)); _auto_width(ws)

    # Hoja: Listado
    ws = wb.create_sheet("Listado")
    ws["A1"] = f"Actualizado al {datetime.today().strftime('%d/%m/%Y')}"
    if not df_listado.empty:
        _write_df(ws, df_listado, start_row=2); _header_style(ws, 2, len(df_listado.columns)); _auto_width(ws)

    # Hoja: Trabajando
    ws = wb.create_sheet("Trabajando")
    ws["A9"]  = "% Según Situación"
    ws["A10"] = "Evolución s/ saldos (+)"
    cols_base = ["Cod","Razón Social","C.Co.","Ej.Cta","P","Límite","Clasificación","Categ.",
                 ">61","46-60","31-45","16-30","1-15","Total general"]
    cols_man  = ["Fecha1","Acción1","Fecha2","Acción2","Interv Ej cta"]
    cols_obs  = [obs_nueva] + ([obs_anterior] if obs_anterior and obs_anterior != obs_nueva else [])
    all_cols  = [c for c in cols_base + [labels["anterior"], labels["actual"]] + cols_man + cols_obs if c in df_trab.columns]
    df_out    = df_trab[all_cols].copy()
    _write_df(ws, df_out, start_row=10)
    _header_style(ws, 10, len(all_cols))
    if obs_nueva in all_cols:
        obs_idx = all_cols.index(obs_nueva) + 1
        yellow  = PatternFill("solid", fgColor="FFFF00")
        for row in range(11, ws.max_row + 1):
            ws.cell(row=row, column=obs_idx).fill = yellow
    _auto_width(ws)

    # Hoja: Base Deuda
    ws = wb.create_sheet("Base Deuda")
    ws["A1"] = "Columna 1"
    ws["S2"] = fecha_ref.date()
    _write_df(ws, df_base_deuda, start_row=3)
    _header_style(ws, 3, len(df_base_deuda.columns))
    _auto_width(ws)

    # Hojas de cobranzas
    # Excel no permite / en nombres de hoja → reemplazar con -
    def _safe_sheet_name(label):
        return label.replace("/", "-")[:31]

    for df_cob, label in [(df_cob_ant, labels["anterior"]), (df_cob_act, labels["actual"])]:
        ws = wb.create_sheet(_safe_sheet_name(label))
        ws["A1"] = "Listado de Movimientos de Ventas"
        if not df_cob.empty:
            _write_df(ws, df_cob, start_row=3)
            _header_style(ws, 3, len(df_cob.columns))
            _auto_width(ws)

    # Hoja: Tabla Dinam
    ws = wb.create_sheet("Tabla Dinam")
    ws["A3"] = "Importes"
    _write_df(ws, df_pivot, start_row=4)
    _header_style(ws, 4, len(df_pivot.columns))
    _auto_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/generar-ad', methods=['POST'])
def endpoint_generar():
    if 'archivo_sistema' not in request.files:
        return jsonify({"error": "archivo_sistema requerido"}), 400

    sistema_bytes  = request.files['archivo_sistema'].read()
    anterior_bytes = request.files['archivo_anterior'].read() if 'archivo_anterior' in request.files else None

    fecha_ref = None
    if request.form.get('fecha_ref'):
        try:
            fecha_ref = datetime.fromisoformat(request.form['fecha_ref'])
        except ValueError:
            return jsonify({"error": "fecha_ref inválida, usar formato YYYY-MM-DD"}), 400

    nombre_salida = request.form.get('nombre_salida', f"AD_{datetime.today().strftime('%d_%m_%Y')}.xlsx")

    try:
        output = _transformar(sistema_bytes, anterior_bytes, fecha_ref)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nombre_salida,
    )


# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5679))
    app.run(host='0.0.0.0', port=port)
